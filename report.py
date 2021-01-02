#!/usr/bin/python

import xmlrpclib
import ssl
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint
import requests
from bs4 import BeautifulSoup

import smtplib
from os.path import basename
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate


#Login details
SATELLITE_URL = "https://vacslp01spacewalk.cucbc.com/rpc/api"
SATELLITE_LOGIN = "Sample_User"
SATELLITE_PASSWORD = "Sample_Pass"
#set ssl checking to off
s=ssl.SSLContext(ssl.PROTOCOL_TLSv1_2)
s.verify_mode=ssl.CERT_NONE

SN_USER = "Sample_User"
SN_PASS = "Sample_Pass"

def get_approver(host):
    try:
        snc_url = "https://central1.service-now.com/cmdb_ci_linux_server_list.do?XML&useUnloadFormat=true&sysparm_query=host_name=%s" % host
        read_url = requests.get(snc_url, auth=(SN_USER, SN_PASS))
        soup = BeautifulSoup(read_url.text, "lxml")
        approver = soup.find("u_access_approved_by")
        approver = str(approver)
        approver = approver.split("\"")
        approver = approver[1]
        if approver == "":
            approver = "Unknown"
    except:
        approver = "Unknown"
    return approver


def lofl_dataset(server):
    base_channel_report = []
    header =("Server", "OS", "OS_Number", "OS_Type", "Base Channel", "Approver")
    base_channel_report.append(header)
    for x in server:
        name = x["name"]
        base = client.system.listBaseChannels(key, x["id"])
        for x in base:
            if x['current_base'] == 1:
                base_channel = x["label"]
                # Set OS Number
                if "-5" in x["label"]:
                    os_number = 5
                elif "-6" in x["label"]:
                    os_number = 6
                elif "-7" in x["label"]:
                    os_number = 7
                elif "-8" in x["label"]:
                    os_number = 8
                else:
                    os_number = 0
                # Set OS
                if "rhel" in x["label"]:
                    os = "Redhat"
                elif "centos" in x["label"]:
                    os = "Centos"
                else:
                    os = "Unknown"
                # Set OS Type
                if "prod" in x["label"]:
                    os_type = "Production"
                elif "qa" in x["label"]:
                    os_type = "Development/QA"
                else:
                    os_type = "Unknown"
                # append all rows
                approver = get_approver(name)
                data = (name, os, os_number, os_type, base_channel, approver)
                base_channel_report.append(data)
    return base_channel_report


def create_spreadsheet(base_channel_report):
    # Setup Excel file
    filename = "server_os_report.xlsx"
    workbook = Workbook()

    ws1 = workbook.create_sheet("Sheet_A")
    ws1.title = "Overview Linux OS"

    ws2 = workbook.create_sheet("Sheet_B")
    ws2.title = "Data"

    ws3 = workbook.create_sheet("Sheet_C")
    ws3.title = "Approver Breakdown"

    sheet = workbook["Data"]

    for row in base_channel_report:
        sheet.append(row)

    darkyellow_background = PatternFill(bgColor=colors.DARKYELLOW)
    yellow_background = PatternFill(bgColor=colors.YELLOW)
    blue_background = PatternFill(bgColor=colors.BLUE)
    green_background = PatternFill(bgColor=colors.GREEN)

    diff_style7 = DifferentialStyle(fill=darkyellow_background)
    rule7 = Rule(type="expression", dxf=diff_style7)
    rule7.formula = ["$C1=7"]
    sheet.conditional_formatting.add("A1:E600", rule7)

    diff_style8 = DifferentialStyle(fill=blue_background)
    rule8 = Rule(type="expression", dxf=diff_style8)
    rule8.formula = ["$C1=7"]
    sheet.conditional_formatting.add("A1:E600", rule8)

    diff_style6 = DifferentialStyle(fill=yellow_background)
    rule6 = Rule(type="expression", dxf=diff_style6)
    rule6.formula = ["$C1=6"]
    sheet.conditional_formatting.add("A1:E600", rule6)

    diff_style5 = DifferentialStyle(fill=green_background)
    rule5 = Rule(type="expression", dxf=diff_style5)
    rule5.formula = ["$C1=5"]
    sheet.conditional_formatting.add("A1:E600", rule5)

    sheet = workbook["Overview Linux OS"]

    data = [
        ['Centos5', '=COUNTIFS(Data!$C$2:$C$600,5, Data!$B$2:$B$600,"Centos")'],
        ['Centos6', '=COUNTIFS(Data!$C$2:$C$600,6, Data!$B$2:$B$600,"Centos")'],
        ['Centos7', '=COUNTIFS(Data!$C$2:$C$600,7, Data!$B$2:$B$600,"Centos")'],
        ['Centos8', '=COUNTIFS(Data!$C$2:$C$600,8, Data!$B$2:$B$600,"Centos")'],
        ['RedHat5', '=COUNTIFS(Data!$C$2:$C$600,5, Data!$B$2:$B$600,"Redhat")'],
        ['RedHat6', '=COUNTIFS(Data!$C$2:$C$600,6, Data!$B$2:$B$600,"Redhat")'],
        ['RedHat7', '=COUNTIFS(Data!$C$2:$C$600,7, Data!$B$2:$B$600,"Redhat")'],
        ['RedHat8', '=COUNTIFS(Data!$C$2:$C$600,8, Data!$B$2:$B$600,"Redhat")'],
        ['Unknown', '=COUNTIFS(Data!$C$2:$C$600,0)']
    ]

    for row in data:
        sheet.append(row)

    pie = PieChart()
    labels = Reference(sheet, min_col=1, min_row=2, max_row=9)
    data = Reference(sheet, min_col=2, min_row=1, max_row=9)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "OS Breakdown"
    pie.height = 20
    pie.width = 40

    # Cut the first slice out of the pie
    slice = DataPoint(idx=0, explosion=20)
    pie.series[0].data_points = [slice]

    sheet.add_chart(pie, "A1")

    std=workbook.get_sheet_by_name('Sheet')
    workbook.remove_sheet(std)


    unique_cost_center = set(x for l in base_channel_report for x in l)

    ws3 = workbook.create_sheet("Sheet_C")
    ws3.title = "Cost Center Breakdown"
    sheet = workbook["Cost Centre Breakdown"]
    data =[]
    for x in unique_cost_center:
        countifs = "=COUNTIFS(Data!$H$2:$H$600,%s)" % x
        data.append([x,countifs])

    for row in data:
        sheet.append(row)

    pie = PieChart()
    labels = Reference(sheet, min_col=1, min_row=2, max_row=len(data))
    data = Reference(sheet, min_col=2, min_row=1, max_row=len(data))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Cost Center Breakdown"
    pie.height = 20
    pie.width = 40

    # Cut the first slice out of the pie
    slice = DataPoint(idx=0, explosion=20)
    pie.series[0].data_points = [slice]

    sheet.add_chart(pie, "A1")
    # save file
    workbook.save(filename)


def send_mail(send_from, send_to, subject, text, files=None,
              server="127.0.0.1"):
    assert isinstance(send_to, list)

    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['To'] = COMMASPACE.join(send_to)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject

    msg.attach(MIMEText(text))

    for f in files or []:
        with open(f, "rb") as fil:
            part = MIMEApplication(
                fil.read(),
                Name=basename(f)
            )
        # After the file is closed
        part['Content-Disposition'] = 'attachment; filename="%s"' % basename(f)
        msg.attach(part)


    smtp = smtplib.SMTP(server)
    smtp.sendmail(send_from, send_to, msg.as_string())
    smtp.close()


if __name__ == "__main__":
    #connetion with ssl checking off
    client = xmlrpclib.Server(SATELLITE_URL, verbose=0,context=s)
    #key for connection
    key = client.auth.login(SATELLITE_LOGIN, SATELLITE_PASSWORD)
    #get sysid based on name which is the first argument against the script
    server = client.system.listSystems(key)
    base_channel_report=lofl_dataset(server)
    create_spreadsheet(base_channel_report)
    #close session
    client.auth.logout(key)
    send_mail("example@example.com", "example@example.com", "Centos 6 EOL Report", "Test Text", files="server_os_report.xlsx")%
